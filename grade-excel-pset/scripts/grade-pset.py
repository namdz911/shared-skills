#!/usr/bin/env python3
"""
Grade Excel problem-set submissions against an answer key.

Auto-detects grading columns in the key, scores all students with 0.1%
relative tolerance, extracts formulas, pattern-matches common errors,
and outputs scoring-report.json.

Usage:
    python3 grade-pset.py \\
        --key path/to/KEY.xlsx \\
        --submissions path/to/submissions/ \\
        --output scoring-report.json

    python3 grade-pset.py \\
        --key path/to/KEY.xlsx \\
        --submissions path/to/submissions/ \\
        --output scoring-report.json \\
        --dry-run
"""

import argparse
import json
import math
import os
import re
import sys

import openpyxl


TOLERANCE = 0.001  # 0.1% relative tolerance
SKIP_SHEETS = {"READ ME", "Grade"}  # Sheets that are not graded


# ---------------------------------------------------------------------------
# Section 1: Key Parsing
# ---------------------------------------------------------------------------

def _find_abs_diff_col(ws):
    """Scan the sheet for a cell containing 'Absolute Difference'. Return (col_letter, row)."""
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "absolute difference" in cell.value.lower():
                return cell.column_letter, cell.row
    return None, None


def _parse_abs_formula(ws_formula, row, diff_col):
    """Parse =ABS(X-Y) formula to get the two column references.

    Handles both simple refs (=ABS(O7-N7)) and cross-sheet refs
    (=ABS('BKE CAPM Output'!B17-T7)) for regression output sheets.
    """
    cell = ws_formula[f"{diff_col}{row}"]
    if not cell.value or not isinstance(cell.value, str):
        return None, None
    formula = cell.value
    # Match =ABS(X1-Y1) pattern (simple same-sheet)
    m = re.match(r"=ABS\(([A-Z]+)\d+\s*-\s*([A-Z]+)\d+\)", formula)
    if m:
        return m.group(1), m.group(2)
    # Match cross-sheet: =ABS('Sheet'!B17-T7) or =ABS(T7-'Sheet'!B17)
    m = re.match(r"=ABS\((?:'[^']*'![A-Z]+\d+|[A-Z]+\d+)\s*-\s*(?:'[^']*'![A-Z]+\d+|[A-Z]+\d+)\)", formula)
    if m:
        # Extract the local column ref (the one without sheet prefix)
        parts = re.findall(r"(?:'[^']*'!)?([A-Z]+)\d+", formula)
        if len(parts) == 2:
            return parts[0], parts[1]
    return None, None


def _identify_key_vs_student(ws_data, col_a, col_b, data_start_row, max_row):
    """Determine which column has the key answers (has values) vs student (empty)."""
    a_has_values = 0
    b_has_values = 0
    for r in range(data_start_row, min(data_start_row + 10, max_row + 1)):
        if ws_data[f"{col_a}{r}"].value is not None:
            a_has_values += 1
        if ws_data[f"{col_b}{r}"].value is not None:
            b_has_values += 1
    if a_has_values > b_has_values:
        return col_a, col_b  # key_col, student_col
    else:
        return col_b, col_a


def _find_graded_rows(ws_formula, diff_col, header_row, max_row):
    """Find all rows that have a grading formula in the diff column."""
    rows = []
    for r in range(header_row + 1, max_row + 1):
        cell = ws_formula[f"{diff_col}{r}"]
        if cell.value and isinstance(cell.value, str) and "ABS(" in cell.value.upper():
            rows.append(r)
    return rows


def _extract_question_text(ws_data, row, key_col):
    """Search columns to the left of the key column for question text in this row or nearby rows."""
    key_col_idx = openpyxl.utils.column_index_from_string(key_col)
    # Look in all columns to the left of the key column for text in this row
    for col_idx in range(key_col_idx - 1, 0, -1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        val = ws_data[f"{col_letter}{row}"].value
        if val and isinstance(val, str) and len(val) > 10:
            return val.strip()
    # If no question in this exact row, search up to 3 rows above
    for offset in range(1, 4):
        r = row - offset
        if r < 1:
            break
        for col_idx in range(key_col_idx - 1, 0, -1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            val = ws_data[f"{col_letter}{r}"].value
            if val and isinstance(val, str) and len(val) > 10:
                return val.strip()
    return ""


def _detect_qualitative_items(wb_data):
    """Detect qualitative items from the Grade sheet.

    The Grade sheet has 'Qualitative Pts (N)' where N is the number of
    qualitative points. These are items graded by the agent (charts, text
    interpretations) rather than by numerical comparison.

    Returns qual_count placeholder items. The actual qualitative grading
    is handled by the agent in a separate pass.
    """
    if "Grade" not in wb_data.sheetnames:
        return []

    ws = wb_data["Grade"]
    # Find "Qualitative Pts (N)" cell to get count
    qual_count = 0
    for row in ws.iter_rows(min_row=1, max_row=25):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                m = re.match(r"Qualitative Pts\s*\((\d+)\)", cell.value)
                if m:
                    qual_count = int(m.group(1))
                    break

    return [
        {
            "item_id": f"qual-{i + 1}",
            "sheet": None,
            "row": None,
            "question": f"Qualitative item {i + 1}",
            "max_points": 1,
            "key_answer": None,
        }
        for i in range(qual_count)
    ]


def _detect_header_row(ws, header_label="Month"):
    """Find the row containing the data column header (e.g., 'Month') in column B.

    Looks for a short cell whose stripped value matches the label exactly,
    to avoid matching substrings in long description cells.
    """
    for r in range(1, 15):
        val = ws.cell(r, 2).value
        if val and isinstance(val, str):
            stripped = val.strip()
            if stripped.lower() == header_label.lower():
                return r
    return None


def _detect_row_offsets(key_path, student_path):
    """Detect per-sheet row offsets between the key and a student file.

    Returns dict of {sheet_name: offset} where offset = key_header_row - student_header_row.
    A positive offset means the key's rows are shifted down relative to the student.
    """
    wb_key = openpyxl.load_workbook(key_path, data_only=True)
    wb_stu = openpyxl.load_workbook(student_path, data_only=True)

    offsets = {}
    for sheet_name in wb_key.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        if sheet_name not in wb_stu.sheetnames:
            # Try stripped match
            matched = None
            for sname in wb_stu.sheetnames:
                if sname.strip() == sheet_name.strip():
                    matched = sname
                    break
            if not matched:
                continue
        else:
            matched = sheet_name

        key_hdr = _detect_header_row(wb_key[sheet_name])
        stu_hdr = _detect_header_row(wb_stu[matched])
        if key_hdr is not None and stu_hdr is not None:
            offsets[sheet_name] = key_hdr - stu_hdr
        else:
            offsets[sheet_name] = 0

    wb_key.close()
    wb_stu.close()
    return offsets


def parse_key(key_path, reference_student_path=None):
    """
    Parse an answer key xlsx to auto-detect grading structure.

    Args:
        key_path: path to the answer key xlsx
        reference_student_path: optional path to a student file for row offset detection.
            If provided, each item gets a 'student_row' field adjusted for any
            header offset between the key and student template.

    Returns:
        dict with keys:
        - "key_file": filename
        - "tolerance": float
        - "items": list of dicts with item_id, sheet, row, student_row, question,
                   correct_value, key_col, student_col
        - "qualitative_items": list of qualitative item dicts
        - "sheets_summary": dict of sheet -> {total_items, key_col, student_col, row_offset}
    """
    # Detect row offsets if a reference student is provided
    row_offsets = {}
    if reference_student_path:
        row_offsets = _detect_row_offsets(key_path, reference_student_path)

    wb_data = openpyxl.load_workbook(key_path, data_only=True)
    wb_formula = openpyxl.load_workbook(key_path, data_only=False)

    items = []
    sheets_summary = {}
    for sheet_name in wb_data.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws_data = wb_data[sheet_name]
        ws_formula = wb_formula[sheet_name]

        # Step 1: Find "Absolute Difference" header
        diff_col, header_row = _find_abs_diff_col(ws_data)
        if diff_col is None:
            # Also check formula sheet (header might be a string value)
            diff_col, header_row = _find_abs_diff_col(ws_formula)
        if diff_col is None:
            continue  # Not a graded sheet

        # Step 2: Find graded rows (rows with ABS formulas)
        graded_rows = _find_graded_rows(ws_formula, diff_col, header_row, ws_data.max_row)
        if not graded_rows:
            continue

        # Step 3: Parse first ABS formula to identify columns
        col_a, col_b = _parse_abs_formula(ws_formula, graded_rows[0], diff_col)
        if col_a is None:
            continue

        # Step 4: Determine which column is key vs student
        key_col, student_col = _identify_key_vs_student(
            ws_data, col_a, col_b, graded_rows[0], ws_data.max_row
        )

        sheet_items = []
        for row in graded_rows:
            correct_value = ws_data[f"{key_col}{row}"].value

            # Skip items where key has no value (shouldn't happen but defensive)
            if correct_value is None:
                continue

            question = _extract_question_text(ws_data, row, key_col)

            # Build a short item_id from sheet name
            sheet_prefix = sheet_name.replace(" ", "")
            offset = row_offsets.get(sheet_name, 0)
            sheet_items.append({
                "item_id": f"{sheet_prefix}-{len(sheet_items) + 1}",
                "sheet": sheet_name,
                "row": row,
                "student_row": row - offset,
                "question": question,
                "correct_value": correct_value,
                "key_col": key_col,
                "student_col": student_col,
            })

        items.extend(sheet_items)
        sheets_summary[sheet_name] = {
            "total_items": len(sheet_items),
            "key_col": key_col,
            "student_col": student_col,
            "diff_col": diff_col,
            "row_offset": row_offsets.get(sheet_name, 0),
        }

    # Detect qualitative items from Grade sheet
    qualitative_items = _detect_qualitative_items(wb_data)

    wb_data.close()
    wb_formula.close()

    return {
        "key_file": os.path.basename(key_path),
        "tolerance": TOLERANCE,
        "items": items,
        "qualitative_items": qualitative_items,
        "sheets_summary": sheets_summary,
    }


# ---------------------------------------------------------------------------
# Section 2: Student Scoring
# ---------------------------------------------------------------------------

def _extract_student_name(filename):
    """Extract student name from Canvas download filename.

    Format: lastfirst_id_timestamp_Description.xlsx
    Example: barnettcole_253721_23009932_FINA363_Excel_Assignment_01 Cole Barnett.xlsx
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    # Try to find name in the latter part of filename (after Canvas IDs)
    parts = base.split("_")
    # Canvas format: login_canvasid_submissionid_rest
    # The "rest" often contains the student's actual name
    if len(parts) >= 4:
        rest = "_".join(parts[3:])
        # Try to extract name from rest (often like "FINA363_Excel_Assignment_01_Cole_Barnett")
        # or "FINA363_Excel_Assignment_01 Cole Barnett"
        # Remove course/assignment prefix patterns
        cleaned = re.sub(r"(?i)FINA\d+[_ ]Excel[_ ]Assign\w*[_ ]\d+[_ ]*", "", rest)
        cleaned = re.sub(r"[-_]+$", "", cleaned)  # trailing separators
        cleaned = cleaned.replace("_", " ").replace("-", " ").strip()
        if cleaned and len(cleaned) > 2:
            return cleaned
    # Fallback: use the login part
    return parts[0] if parts else base


def _scan_student_sheet_for_value(ws, key_row, correct_value, row_range=5, min_col=4, prefer_col=None):
    """Scan the student sheet near key_row for a cell matching correct_value.

    The student template may have a different row layout or column structure
    than the key (the key adds extra grading columns). This scanner locates the
    student's answer by searching nearby cells for a value within TOLERANCE.

    Args:
        ws: student worksheet (data_only)
        key_row: row number from the answer key
        correct_value: expected numeric value
        row_range: rows to scan above/below key_row
        min_col: skip data-heavy left columns (raw time-series data)
        prefer_col: column index to prefer when multiple candidates match.
            If provided, returns the candidate nearest to this column.

    Returns:
        (is_correct, student_value) tuple
    """
    if correct_value is None:
        return False, None
    try:
        cv = float(correct_value)
    except (ValueError, TypeError):
        return False, None

    tol = abs(cv) * TOLERANCE if cv != 0 else 1e-9

    # Collect all matching candidates: (row_distance, col_distance, value)
    candidates = []
    for r_offset in range(-row_range, row_range + 1):
        sr = key_row + r_offset
        if sr < 1 or sr > ws.max_row:
            continue
        for c in range(min_col, ws.max_column + 1):
            sv = ws.cell(sr, c).value
            if sv is not None and isinstance(sv, (int, float)):
                try:
                    if abs(float(sv) - cv) <= tol:
                        col_dist = abs(c - prefer_col) if prefer_col else 0
                        candidates.append((abs(r_offset), col_dist, float(sv)))
                except (ValueError, TypeError):
                    pass

    if not candidates:
        return False, None

    # Sort by: column distance to preferred col, then row distance
    candidates.sort(key=lambda x: (x[1], x[0]))
    return True, candidates[0][2]


def score_student(student_path, grading_map):
    """
    Score a student's submission against the grading map.

    Uses a scan-based approach: for each graded item, searches nearby cells
    in the student sheet for a value matching the key answer within TOLERANCE.
    This handles differences in column layout and minor row offsets between
    the key template and student submission.

    Args:
        student_path: path to student xlsx file
        grading_map: output from parse_key()

    Returns:
        dict with name, file, scores (per-sheet), total_numerical,
        total_possible, items (list of per-item results)
    """
    try:
        wb = openpyxl.load_workbook(student_path, data_only=True)
    except Exception as e:
        return {
            "name": _extract_student_name(student_path),
            "file": os.path.basename(student_path),
            "error": f"Could not open file: {e}",
            "scores": {},
            "total_numerical": 0,
            "total_possible": len(grading_map["items"]),
            "items": [],
        }

    student_name = _extract_student_name(student_path)
    per_sheet_scores = {}
    item_results = []

    # Cache open worksheets per sheet name
    ws_cache = {}

    for item in grading_map["items"]:
        sheet = item["sheet"]
        row = item.get("student_row", item["row"])  # Use student_row if available
        correct_value = item["correct_value"]

        # Initialize per-sheet counters
        if sheet not in per_sheet_scores:
            per_sheet_scores[sheet] = {"right": 0, "total": 0}
        per_sheet_scores[sheet]["total"] += 1

        # Find the sheet in student file (handle trailing-space variants)
        ws = ws_cache.get(sheet)
        if ws is None:
            matched_sheet = None
            if sheet in wb.sheetnames:
                matched_sheet = sheet
            else:
                # Try stripped names in case of trailing spaces
                for sname in wb.sheetnames:
                    if sname.strip() == sheet.strip():
                        matched_sheet = sname
                        break
            if matched_sheet:
                ws_cache[sheet] = wb[matched_sheet]
                ws = ws_cache[sheet]

        is_correct = False
        student_value = None

        if ws is not None:
            # Prefer the expected student answer column when multiple candidates match
            prefer_col_idx = None
            student_col = item.get("student_col")
            if student_col:
                try:
                    prefer_col_idx = openpyxl.utils.column_index_from_string(student_col)
                except ValueError:
                    pass
            is_correct, student_value = _scan_student_sheet_for_value(
                ws, row, correct_value, prefer_col=prefer_col_idx
            )

        if is_correct:
            per_sheet_scores[sheet]["right"] += 1

        result = {
            "item_id": item["item_id"],
            "sheet": sheet,
            "row": item["row"],
            "student_row": row,
            "correct": is_correct,
            "student_value": student_value,
            "correct_value": correct_value,
            "question": item["question"],
        }

        # Only include wrong items in detail (keeps JSON smaller)
        if not is_correct:
            item_results.append(result)

    wb.close()

    total_right = sum(s["right"] for s in per_sheet_scores.values())
    total_possible = sum(s["total"] for s in per_sheet_scores.values())

    return {
        "name": student_name,
        "file": os.path.basename(student_path),
        "scores": per_sheet_scores,
        "total_numerical": total_right,
        "total_possible": total_possible,
        "percentage": total_right / total_possible if total_possible > 0 else 0,
        "items": item_results,
    }


# ---------------------------------------------------------------------------
# Section 3: Formula Extraction
# ---------------------------------------------------------------------------

def _cell_formula_str(cell_value):
    """Normalize a cell value to a formula string.

    Handles plain strings (formulas like '=AVERAGE(...)'), ArrayFormula
    objects (openpyxl wraps CSE/array formulas), and literal values.

    Returns a string or None.
    """
    if cell_value is None:
        return None
    # ArrayFormula objects have a .text attribute with the formula text
    if hasattr(cell_value, "text"):
        text = cell_value.text
        if text and not text.startswith("="):
            text = "=" + text
        return text
    if isinstance(cell_value, str):
        return cell_value
    # Numeric or date literal
    return str(cell_value)


def extract_formulas(student_path, grading_map):
    """Extract formulas from a student's submission for all graded items.

    Opens the file with data_only=False to read cell formulas, and also
    with data_only=True to verify the computed value.  For each graded item,
    scans near the expected row for the student's answer cell.

    Because student files have different column layouts than the key
    (the key adds extra grading columns), we use the same scan-based approach
    as score_student() rather than trusting the key's student_col field.

    Args:
        student_path: path to student xlsx file
        grading_map: output from parse_key()

    Returns:
        list of dicts: {item_id, sheet, row, cell, formula, is_formula}
        - cell: cell address like 'J7' (None if not found)
        - formula: formula string like '=AVERAGE(E5:E1168)' or literal '0.0065'
        - is_formula: True if the cell contained an Excel formula
    """
    try:
        wb_formula = openpyxl.load_workbook(student_path, data_only=False)
        wb_data = openpyxl.load_workbook(student_path, data_only=True)
    except Exception as e:
        return []

    results = []
    ws_formula_cache = {}
    ws_data_cache = {}

    for item in grading_map["items"]:
        sheet = item["sheet"]
        row = item.get("student_row", item["row"])  # Use student_row if available
        correct_value = item["correct_value"]

        # Resolve worksheet — handle trailing-space name variants
        ws_f = ws_formula_cache.get(sheet)
        ws_d = ws_data_cache.get(sheet)
        if ws_f is None:
            matched = None
            if sheet in wb_formula.sheetnames:
                matched = sheet
            else:
                for sname in wb_formula.sheetnames:
                    if sname.strip() == sheet.strip():
                        matched = sname
                        break
            if matched:
                ws_formula_cache[sheet] = wb_formula[matched]
                ws_data_cache[sheet] = wb_data[matched]
                ws_f = ws_formula_cache[sheet]
                ws_d = ws_data_cache[sheet]

        cell_addr = None
        formula_str = None
        is_formula = False

        if ws_f is not None:
            # First pass: look for a formula cell at a location whose DATA
            # value matches the correct answer (best case — found the right cell)
            try:
                cv = float(correct_value)
                tol = abs(cv) * TOLERANCE if cv != 0 else 1e-9
            except (ValueError, TypeError):
                cv = None
                tol = 0

            # Determine preferred column for this item
            prefer_col_idx_f = None
            student_col_f = item.get("student_col")
            if student_col_f:
                try:
                    prefer_col_idx_f = openpyxl.utils.column_index_from_string(student_col_f)
                except ValueError:
                    pass

            found_exact = False
            pass1_candidates = []
            for r_offset in range(-5, 6):
                sr = row + r_offset
                if sr < 1 or sr > ws_f.max_row:
                    continue
                for c in range(4, ws_f.max_column + 1):
                    cell_f = ws_f.cell(sr, c)
                    fstr = _cell_formula_str(cell_f.value)
                    if fstr is None or "FORMULATEXT" in (fstr.upper() if isinstance(fstr, str) else ""):
                        continue
                    if fstr.startswith("="):
                        # Verify data value if possible
                        if cv is not None and ws_d is not None:
                            dv = ws_d.cell(sr, c).value
                            if dv is not None and isinstance(dv, (int, float)):
                                if abs(float(dv) - cv) <= tol:
                                    col_dist = abs(c - prefer_col_idx_f) if prefer_col_idx_f else 0
                                    pass1_candidates.append((abs(r_offset), col_dist, c, sr, fstr))

            if pass1_candidates:
                # Prefer nearest to expected student_col, then nearest row
                pass1_candidates.sort(key=lambda x: (x[1], x[0]))
                _, _, best_c, best_sr, best_fstr = pass1_candidates[0]
                col_letter = openpyxl.utils.get_column_letter(best_c)
                cell_addr = f"{col_letter}{best_sr}"
                formula_str = best_fstr
                is_formula = True
                found_exact = True

            if not found_exact:
                # Second pass: look for any numeric literal matching correct_value
                pass2_candidates = []
                for r_offset in range(-5, 6):
                    sr = row + r_offset
                    if sr < 1 or sr > ws_f.max_row:
                        continue
                    for c in range(4, ws_f.max_column + 1):
                        cell_f = ws_f.cell(sr, c)
                        fstr = _cell_formula_str(cell_f.value)
                        if fstr is None or fstr.startswith("="):
                            continue
                        # Try to match as numeric literal
                        try:
                            if cv is not None and abs(float(fstr) - cv) <= tol:
                                col_dist = abs(c - prefer_col_idx_f) if prefer_col_idx_f else 0
                                pass2_candidates.append((abs(r_offset), col_dist, c, sr, fstr))
                        except (ValueError, TypeError):
                            pass

                if pass2_candidates:
                    pass2_candidates.sort(key=lambda x: (x[1], x[0]))
                    _, _, best_c, best_sr, best_fstr = pass2_candidates[0]
                    col_letter = openpyxl.utils.get_column_letter(best_c)
                    cell_addr = f"{col_letter}{best_sr}"
                    formula_str = best_fstr
                    is_formula = False
                    found_exact = True

            if not found_exact:
                # Third pass: extract formula at the expected location even for
                # wrong answers — used for pattern diagnosis.  Collect all
                # formula cells (not FORMULATEXT) in the scan window, then
                # prefer the one nearest to the expected student_col.
                prefer_col_idx = None
                student_col = item.get("student_col")
                if student_col:
                    try:
                        prefer_col_idx = openpyxl.utils.column_index_from_string(student_col)
                    except ValueError:
                        pass

                formula_candidates = []
                for r_offset in range(0, 6):
                    for sign in (1, -1) if r_offset != 0 else (1,):
                        sr = row + r_offset * sign
                        if sr < 1 or sr > ws_f.max_row:
                            continue
                        for c in range(4, ws_f.max_column + 1):
                            cell_f = ws_f.cell(sr, c)
                            fstr = _cell_formula_str(cell_f.value)
                            if fstr is None:
                                continue
                            if fstr.startswith("=") and "FORMULATEXT" not in fstr.upper():
                                col_dist = abs(c - prefer_col_idx) if prefer_col_idx else 0
                                formula_candidates.append((abs(r_offset), col_dist, c, sr, fstr))

                if formula_candidates:
                    # Sort by: column distance to expected, then row distance
                    formula_candidates.sort(key=lambda x: (x[1], x[0]))
                    _, _, best_c, best_sr, best_fstr = formula_candidates[0]
                    col_letter = openpyxl.utils.get_column_letter(best_c)
                    cell_addr = f"{col_letter}{best_sr}"
                    formula_str = best_fstr
                    is_formula = True
                    found_exact = True

        # Also capture the data value at the found cell (for wrong-answer diagnosis)
        actual_value = None
        if cell_addr and ws_d is not None:
            try:
                col_letter, cell_row = openpyxl.utils.cell.coordinate_from_string(cell_addr)
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                dv = ws_d.cell(cell_row, col_idx).value
                if isinstance(dv, (int, float)):
                    actual_value = float(dv)
            except Exception:
                pass

        results.append({
            "item_id": item["item_id"],
            "sheet": sheet,
            "row": row,
            "cell": cell_addr,
            "formula": formula_str,
            "is_formula": is_formula,
            "actual_value": actual_value,
        })

    wb_formula.close()
    wb_data.close()

    return results


# ---------------------------------------------------------------------------
# Section 4: Pattern Matching
# ---------------------------------------------------------------------------


def _patterns_for_item(item_result, grading_map, formula_entry):
    """Identify error patterns for a single wrong item.

    Args:
        item_result: one element from score_student()["items"] (wrong item)
        grading_map: parse_key() output
        formula_entry: matching entry from extract_formulas() or None

    Returns:
        (pattern_match, pattern_detail) strings or (None, None)
    """
    correct_value = item_result.get("correct_value")
    student_value = item_result.get("student_value")
    formula = formula_entry.get("formula") if formula_entry else None

    # Use the actual_value from formula extraction when score_student returned None
    # (score_student only populates student_value when the answer is correct)
    if student_value is None and formula_entry:
        student_value = formula_entry.get("actual_value")

    # Pattern 1: missing_answer — blank or zero
    if student_value is None or student_value == 0:
        return "missing_answer", "cell is blank or zero"

    try:
        cv = float(correct_value)
        sv = float(student_value)
    except (ValueError, TypeError):
        return None, None

    # Pattern 2: sign_error — student ≈ -correct
    if cv != 0:
        tol = abs(cv) * TOLERANCE
        if abs(sv - (-cv)) <= tol:
            return "sign_error", f"student={sv:.6g}, correct={cv:.6g} (sign flipped)"

    # Pattern 3: gave_monthly_not_annual — student ≈ correct/12 or correct/sqrt(12)
    if cv != 0:
        tol = abs(cv) * 0.01  # 1% tolerance for this heuristic
        if abs(sv - cv / 12) <= tol:
            return "gave_monthly_not_annual", f"student={sv:.6g} ≈ correct/12={cv/12:.6g}"
        if abs(sv - cv / math.sqrt(12)) <= tol:
            return "gave_monthly_not_annual", f"student={sv:.6g} ≈ correct/sqrt(12)={cv/math.sqrt(12):.6g}"

    # Pattern 4: annualization_error — multiplied by 12 instead of compounding
    # Detect: student used *12 when they should compound (1+r)^12-1
    # The annualized return via compounding is (1+r_monthly)^12-1
    # The simple multiplication gives r_monthly * 12
    # If the correct answer looks like a compounded return and student gave *12 version
    if formula and isinstance(formula, str):
        formula_upper = formula.upper()

        # Pattern 5: used_population_stat — STDEV.P or VAR.P instead of sample stat
        if re.search(r"STDEV\.P\b|STDEVP\b|VAR\.P\b|VARP\b", formula_upper):
            return "used_population_stat", f"formula uses population statistic: {formula}"

        # Pattern 6: wrong_range — formula references a truncated range
        # Heuristic: look for two similar range functions in the same formula
        # where one has a significantly shorter range, or compare start rows.
        # Conservative: only flag ranges that start at the same row as
        # another range in the formula but end much earlier.
        ranges_in_formula = re.findall(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", formula_upper)
        if len(ranges_in_formula) >= 2:
            # Multiple ranges — check if any is significantly shorter
            row_spans = [(int(r[1]), int(r[3]), int(r[3]) - int(r[1]) + 1) for r in ranges_in_formula]
            max_span = max(s[2] for s in row_spans)
            for start, end, span in row_spans:
                if span < max_span * 0.5 and max_span > 20:
                    return "wrong_range", f"formula range ends at row {end} vs max range of {max_span} rows: {formula}"
        elif len(ranges_in_formula) == 1:
            start_row = int(ranges_in_formula[0][1])
            end_row = int(ranges_in_formula[0][3])
            span = end_row - start_row + 1
            # Only flag single ranges that look clearly truncated:
            # started at a typical data start row but ended too early
            if start_row < 10 and 20 < span < 500 and end_row < 900:
                return "wrong_range", f"formula range {ranges_in_formula[0][0]}{start_row}:{ranges_in_formula[0][2]}{end_row} covers {span} rows — may be truncated: {formula}"

        # Pattern 7: annualization_error — *12 instead of compounding
        if re.search(r"\*\s*12\b", formula_upper) and not re.search(r"SQRT\s*\(", formula_upper):
            # Check if the correct answer looks like a compounded value
            # If sv * (1/12) is close to something that when compounded gives cv
            # i.e., (1 + sv/12)^12 - 1 ≈ cv
            if cv != 0 and abs(cv) < 5:  # reasonable return value
                monthly_implied = sv / 12
                if monthly_implied > -1:
                    compounded = (1 + monthly_implied) ** 12 - 1
                    if abs(compounded - cv) / max(abs(cv), 1e-9) < 0.05:
                        return "annualization_error", f"multiplied by 12 instead of compounding; formula: {formula}"

    # Pattern: wrong_item_value — student answer matches a different item's correct value
    all_correct = [
        float(it["correct_value"])
        for it in grading_map["items"]
        if it["correct_value"] is not None and it["item_id"] != item_result["item_id"]
    ]
    if cv != 0:
        tol = abs(cv) * 0.01  # 1% for cross-item match
        for other_cv in all_correct:
            if abs(sv - other_cv) / max(abs(other_cv), 1e-9) < 0.01:
                return "wrong_item_value", f"student={sv:.6g} matches another item's answer={other_cv:.6g}"

    # Pattern 2b: hardcoded — literal value instead of formula (check after others)
    if formula_entry and formula_entry.get("formula") and not formula_entry.get("is_formula"):
        return "hardcoded", f"literal value {formula_entry['formula']} instead of formula"

    return None, None


def match_patterns(student_result, grading_map, formulas):
    """For each wrong answer, try to identify common error patterns.

    Checks 7 patterns: missing_answer, hardcoded, sign_error,
    gave_monthly_not_annual, used_population_stat, wrong_range,
    annualization_error; plus wrong_item_value.

    Args:
        student_result: output from score_student()
        grading_map: output from parse_key()
        formulas: output from extract_formulas()

    Returns:
        list of dicts with item_id, pattern_match, pattern_detail,
        student_formula for each wrong item.
    """
    # Build formula lookup by item_id
    formula_by_item = {f["item_id"]: f for f in formulas}

    pattern_results = []
    for item in student_result.get("items", []):
        item_id = item["item_id"]
        formula_entry = formula_by_item.get(item_id)

        pattern_match, pattern_detail = _patterns_for_item(
            item, grading_map, formula_entry
        )

        pattern_results.append({
            "item_id": item_id,
            "sheet": item.get("sheet"),
            "row": item.get("row"),
            "pattern_match": pattern_match,
            "pattern_detail": pattern_detail,
            "student_value": item.get("student_value"),
            "correct_value": item.get("correct_value"),
            "student_formula": formula_entry.get("formula") if formula_entry else None,
        })

    return pattern_results


# ---------------------------------------------------------------------------
# Section 5: Batch Scoring + JSON Output
# ---------------------------------------------------------------------------

def score_all_students(key_path, submissions_dir, grading_map):
    """
    Score all student submissions in a directory.
    Returns the full scoring report dict ready for JSON serialization.
    """
    # Find all xlsx files in submissions directory
    student_files = []
    for fname in sorted(os.listdir(submissions_dir)):
        if fname.lower().endswith((".xlsx", ".xls")) and not fname.startswith("~"):
            student_files.append(os.path.join(submissions_dir, fname))

    print(f"Found {len(student_files)} student submissions")

    students = []
    for i, student_path in enumerate(student_files, 1):
        name = os.path.basename(student_path)
        print(f"  [{i}/{len(student_files)}] {name[:60]}...", end=" ", flush=True)

        # Score
        result = score_student(student_path, grading_map)

        # Extract formulas for wrong items
        if result["items"]:  # has wrong answers
            formulas = extract_formulas(student_path, grading_map)
            patterns = match_patterns(result, grading_map, formulas)
            # Merge pattern + cell info into wrong items
            pattern_lookup = {p["item_id"]: p for p in patterns}
            formula_lookup = {f["item_id"]: f for f in formulas}
            for item in result["items"]:
                p = pattern_lookup.get(item["item_id"], {})
                f = formula_lookup.get(item["item_id"], {})
                item["pattern_match"] = p.get("pattern_match")
                item["pattern_detail"] = p.get("pattern_detail")
                item["student_formula"] = p.get("student_formula")
                item["student_cell"] = f.get("cell")
                if item.get("student_value") is None and f.get("actual_value") is not None:
                    item["student_value"] = f["actual_value"]

        pct = result["percentage"]
        wrong = len(result["items"])
        print(f"{result['total_numerical']}/{result['total_possible']} ({pct:.0%}) — {wrong} wrong")

        students.append(result)

    # Build grading map for output (simplified)
    gmap_output = []
    for item in grading_map["items"]:
        gmap_output.append({
            "item_id": item["item_id"],
            "sheet": item["sheet"],
            "row": item["row"],
            "student_row": item.get("student_row", item["row"]),
            "question": item["question"],
            "correct_value": item["correct_value"],
            "key_col": item["key_col"],
            "student_col": item["student_col"],
        })

    # Sort students by name
    students.sort(key=lambda s: s["name"].lower())

    assignment_name = os.path.splitext(grading_map["key_file"])[0]

    return {
        "assignment": assignment_name,
        "key_file": grading_map["key_file"],
        "tolerance": grading_map["tolerance"],
        "total_students": len(students),
        "grading_map": gmap_output,
        "qualitative_items": grading_map["qualitative_items"],
        "sheets_summary": grading_map["sheets_summary"],
        "students": students,
    }


def write_report(report, output_path):
    """Write scoring report to JSON file."""
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"\nReport written to: {output_path}")
    print(f"  {report['total_students']} students")
    print(f"  {len(report['grading_map'])} graded items")

    # Print summary stats
    pcts = [s["percentage"] for s in report["students"]]
    if pcts:
        avg = sum(pcts) / len(pcts)
        lo = min(pcts)
        hi = max(pcts)
        print(f"  Scores: avg {avg:.0%}, min {lo:.0%}, max {hi:.0%}")


# ---------------------------------------------------------------------------
# Section 6: CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Grade Excel problem-set submissions against an answer key",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--key", required=True,
                        help="Path to the answer key .xlsx file")
    parser.add_argument("--submissions", required=True,
                        help="Path to directory containing student .xlsx files")
    parser.add_argument("--output", default="scoring-report.json",
                        help="Output path for scoring report (default: scoring-report.json)")
    parser.add_argument("--workspace",
                        help="Base workspace directory (per script-conventions.md). "
                             "If set, --key and --submissions are relative to this path.")
    parser.add_argument("--reference-student",
                        help="Path to a student file for row offset detection. "
                             "If not provided, auto-selects the first submission.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse key and count submissions without scoring")

    args = parser.parse_args()

    # Resolve paths relative to workspace if provided
    if args.workspace:
        if not os.path.isdir(args.workspace):
            print(f"Error: Workspace directory not found: {args.workspace}")
            sys.exit(1)
        args.key = os.path.join(args.workspace, args.key)
        args.submissions = os.path.join(args.workspace, args.submissions)
        if args.output == "scoring-report.json":
            args.output = os.path.join(args.workspace, args.output)

    # Validate inputs
    if not os.path.isfile(args.key):
        print(f"Error: Key file not found: {args.key}")
        sys.exit(1)
    if not os.path.isdir(args.submissions):
        print(f"Error: Submissions directory not found: {args.submissions}")
        sys.exit(1)

    # Auto-detect reference student if not provided
    ref_student = args.reference_student
    if not ref_student:
        sub_files = sorted([
            os.path.join(args.submissions, f)
            for f in os.listdir(args.submissions)
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~")
        ])
        if sub_files:
            ref_student = sub_files[0]

    # Parse key
    print(f"Parsing key: {args.key}")
    if ref_student:
        print(f"  Reference student for offset detection: {os.path.basename(ref_student)}")
    grading_map = parse_key(args.key, reference_student_path=ref_student)
    print(f"  Detected {len(grading_map['items'])} graded items across "
          f"{len(grading_map['sheets_summary'])} sheets")
    for sheet, info in grading_map["sheets_summary"].items():
        offset = info.get("row_offset", 0)
        offset_str = f", row_offset={offset}" if offset != 0 else ""
        print(f"    {sheet}: {info['total_items']} items "
              f"(key={info['key_col']}, student={info['student_col']}{offset_str})")

    # Count submissions
    sub_files = [f for f in os.listdir(args.submissions)
                 if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~")]
    print(f"\nFound {len(sub_files)} student submissions in {args.submissions}")

    if args.dry_run:
        print("\n[DRY RUN] No scoring performed.")
        return

    # Score all students
    print(f"\nScoring...")
    report = score_all_students(args.key, args.submissions, grading_map)

    # Write output
    write_report(report, args.output)


if __name__ == "__main__":
    main()
